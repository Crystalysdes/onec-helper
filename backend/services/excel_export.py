"""Excel export service for Kontur Market import format."""
import io
from typing import List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_KM_COLUMNS = [
    ("Наименование", 40),
    ("Штрихкод", 18),
    ("Артикул", 16),
    ("Единица измерения", 14),
    ("Группа (категория)", 20),
    ("Закупочная цена", 14),
    ("Розничная цена", 14),
    ("НДС %", 8),
    ("Количество", 12),
    ("Описание", 30),
]

_HEADER_FILL = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
_ALT_FILL    = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
_BORDER_SIDE = Side(style="thin", color="CBD5E1")
_THIN_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE,
)


def _hcell(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _THIN_BORDER


def _dcell(ws, row: int, col: int, value, alt: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    c.border = _THIN_BORDER
    if alt:
        c.fill = _ALT_FILL
    return c


def generate_kontur_market_xlsx(
    products: List[dict],
    store_name: Optional[str] = None,
) -> bytes:
    """Generate XLSX file in Kontur Market import format.

    Args:
        products: list of product dicts with keys:
            name, barcode, article, unit, category,
            purchase_price, price, quantity, description
        store_name: optional store name for the sheet title

    Returns:
        bytes of the XLSX file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    # ── Info row ──────────────────────────────────────────────────
    title = f"Импорт товаров для Контур.Маркет"
    if store_name:
        title += f" — {store_name}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_KM_COLUMNS))
    info = ws.cell(row=1, column=1, value=title)
    info.font = Font(bold=True, size=12, color="1A56DB")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_KM_COLUMNS))
    dt_cell = ws.cell(
        row=2, column=1,
        value=f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Товаров: {len(products)}"
    )
    dt_cell.font = Font(size=9, color="64748B")
    dt_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Header row ────────────────────────────────────────────────
    HEADER_ROW = 3
    ws.row_dimensions[HEADER_ROW].height = 36
    for col_idx, (col_name, col_width) in enumerate(_KM_COLUMNS, start=1):
        _hcell(ws, HEADER_ROW, col_idx, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, p in enumerate(products, start=HEADER_ROW + 1):
        alt = (row_idx % 2 == 0)
        ws.row_dimensions[row_idx].height = 18

        name = (p.get("name") or "").strip()
        barcode = p.get("barcode") or ""
        article = p.get("article") or ""
        unit = p.get("unit") or "шт"
        category = p.get("category") or ""
        purchase_price = p.get("purchase_price")
        price = p.get("price")
        quantity = p.get("quantity") or 0
        description = p.get("description") or ""

        _dcell(ws, row_idx, 1, name, alt)
        bc = _dcell(ws, row_idx, 2, str(barcode) if barcode else "", alt)
        bc.number_format = "@"  # keep as text to preserve leading zeros
        _dcell(ws, row_idx, 3, str(article) if article else "", alt)
        _dcell(ws, row_idx, 4, unit, alt)
        _dcell(ws, row_idx, 5, category, alt)

        pp = _dcell(ws, row_idx, 6, round(float(purchase_price), 2) if purchase_price is not None else "", alt)
        if purchase_price is not None:
            pp.number_format = '#,##0.00'

        rp = _dcell(ws, row_idx, 7, round(float(price), 2) if price is not None else "", alt)
        if price is not None:
            rp.number_format = '#,##0.00'

        _dcell(ws, row_idx, 8, 20, alt)  # НДС 20% default

        qty = _dcell(ws, row_idx, 9, round(float(quantity), 3) if quantity else 0, alt)
        qty.number_format = '#,##0.###'

        _dcell(ws, row_idx, 10, description, alt)

    # ── Freeze header ─────────────────────────────────────────────
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── Auto-filter ───────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{HEADER_ROW}:{get_column_letter(len(_KM_COLUMNS))}{HEADER_ROW + len(products)}"
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_kontur_market_csv(products: List[dict]) -> bytes:
    """Lightweight CSV fallback (semicolon-delimited, Windows-1251)."""
    rows = ["Наименование;Штрихкод;Артикул;Ед.изм;Группа;Закупочная цена;Розничная цена;НДС%;Количество"]
    for p in products:
        def _s(v): return str(v).replace(";", ",") if v is not None else ""
        rows.append(";".join([
            _s(p.get("name", "")),
            _s(p.get("barcode", "")),
            _s(p.get("article", "")),
            _s(p.get("unit", "шт")),
            _s(p.get("category", "")),
            _s(p.get("purchase_price", "")),
            _s(p.get("price", "")),
            "20",
            _s(p.get("quantity", 0)),
        ]))
    return "\n".join(rows).encode("utf-8-sig")  # BOM for Excel compatibility
