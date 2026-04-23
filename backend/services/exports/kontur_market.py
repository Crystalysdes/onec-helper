"""Kontur.Market import format.

Column order matches the official import template from market.kontur.ru:
    A: Наименование товара
    B: Штрихкод
    C: Розничная цена
    D: Ед. измерения
    E: Артикул
    F: Ставка НДС
    G: Тип товара
    H: Остаток
    I: Закупочная цена
    J: Группа товара
    K: Группа 2 уровня
    L: Группа 3 уровня

Row 1 is the header; data starts from row 2. No merged cells or styling
rows — Kontur's importer expects a flat table.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .base import ExportFormat, ProductDict


_COLUMNS = [
    ("Наименование товара", 40),
    ("Штрихкод",            18),
    ("Розничная цена",      14),
    ("Ед. измерения",       14),
    ("Артикул",             16),
    ("Ставка НДС",          10),
    ("Тип товара",          14),
    ("Остаток",             12),
    ("Закупочная цена",     14),
    ("Группа товара",       20),
    ("Группа 2 уровня",     20),
    ("Группа 3 уровня",     20),
]


class KonturMarketFormat(ExportFormat):
    format_id = "kontur_market"
    label = "Контур.Маркет"
    description = "Импорт товаров в Контур.Маркет (официальный шаблон)"
    target = "kontur"

    def generate(self, products: List[ProductDict], meta: Dict[str, Any] | None = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        # Header row (row 1)
        for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, p in enumerate(products, start=2):
            name = self.safe_str(p.get("name"))
            barcode = self.safe_str(p.get("barcode"))
            price = self.safe_num(p.get("price"))
            unit = self.safe_str(p.get("unit")) or "шт"
            article = self.safe_str(p.get("article"))
            vat = p.get("vat_rate")
            vat_str = self._format_vat(vat)
            product_type = self.safe_str(p.get("product_type")) or "Товар"
            qty = self.safe_num(p.get("quantity"), default=0)
            purchase = self.safe_num(p.get("purchase_price"))
            cat = self.safe_str(p.get("category"))
            cat2 = self.safe_str(p.get("category_l2"))
            cat3 = self.safe_str(p.get("category_l3"))

            ws.cell(row=row_idx, column=1, value=name)

            bc = ws.cell(row=row_idx, column=2, value=barcode)
            bc.number_format = "@"  # preserve leading zeros

            if price is not None:
                pc = ws.cell(row=row_idx, column=3, value=round(price, 2))
                pc.number_format = "0.00"

            ws.cell(row=row_idx, column=4, value=unit)
            ws.cell(row=row_idx, column=5, value=article)
            ws.cell(row=row_idx, column=6, value=vat_str)
            ws.cell(row=row_idx, column=7, value=product_type)

            if qty is not None:
                qc = ws.cell(row=row_idx, column=8, value=round(qty, 3))
                qc.number_format = "0.###"

            if purchase is not None:
                pp = ws.cell(row=row_idx, column=9, value=round(purchase, 2))
                pp.number_format = "0.00"

            ws.cell(row=row_idx, column=10, value=cat)
            ws.cell(row=row_idx, column=11, value=cat2)
            ws.cell(row=row_idx, column=12, value=cat3)

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def _format_vat(value: Any) -> str:
        """Kontur expects the VAT rate as a percent string ("0%", "10%", "20%") or "Без НДС"."""
        if value is None or value == "":
            return "20%"
        try:
            n = int(float(value))
        except (TypeError, ValueError):
            s = str(value).strip()
            return s if s else "20%"
        if n <= 0:
            return "Без НДС"
        return f"{n}%"
