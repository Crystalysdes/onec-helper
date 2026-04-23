"""Export format for 1С:Розница (standard Excel catalog import).

1С:Розница and other БСП-based configs use the built-in "Загрузка данных
из файла Excel" wizard (НСИ — Сервис) which lets the user map any column
to any attribute. We emit a stable column set with names that auto-match
on the standard wizard so no manual mapping is required.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .base import ExportFormat, ProductDict


_COLUMNS = [
    ("Наименование",        40),
    ("Артикул",             16),
    ("Штрихкод",            18),
    ("Единица измерения",   14),
    ("Вид номенклатуры",    16),
    ("Группа",              22),
    ("Ставка НДС",          10),
    ("Цена",                14),
    ("Закупочная цена",     14),
    ("Остаток",             12),
]


class OneCRetailFormat(ExportFormat):
    format_id = "onec_retail"
    label = "1С:Розница"
    description = "Импорт номенклатуры в 1С:Розница (Сервис «Загрузка данных из Excel»)"
    target = "onec"

    def generate(self, products: List[ProductDict], meta: Dict[str, Any] | None = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Номенклатура"

        for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, p in enumerate(products, start=2):
            name = self.safe_str(p.get("name"))
            article = self.safe_str(p.get("article"))
            barcode = self.safe_str(p.get("barcode"))
            unit = self.safe_str(p.get("unit")) or "шт"
            product_type = self.safe_str(p.get("product_type")) or "Товар"
            category = self.safe_str(p.get("category"))
            vat = p.get("vat_rate")
            vat_str = self._format_vat(vat)
            price = self.safe_num(p.get("price"))
            purchase = self.safe_num(p.get("purchase_price"))
            qty = self.safe_num(p.get("quantity"), default=0)

            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=article)
            bc = ws.cell(row=row_idx, column=3, value=barcode)
            bc.number_format = "@"
            ws.cell(row=row_idx, column=4, value=unit)
            ws.cell(row=row_idx, column=5, value=product_type)
            ws.cell(row=row_idx, column=6, value=category)
            ws.cell(row=row_idx, column=7, value=vat_str)
            if price is not None:
                c = ws.cell(row=row_idx, column=8, value=round(price, 2))
                c.number_format = "0.00"
            if purchase is not None:
                c = ws.cell(row=row_idx, column=9, value=round(purchase, 2))
                c.number_format = "0.00"
            if qty is not None:
                c = ws.cell(row=row_idx, column=10, value=round(qty, 3))
                c.number_format = "0.###"

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def _format_vat(value: Any) -> str:
        if value is None or value == "":
            return "НДС 20%"
        try:
            n = int(float(value))
        except (TypeError, ValueError):
            s = str(value).strip()
            return s or "НДС 20%"
        if n <= 0:
            return "Без НДС"
        return f"НДС {n}%"
