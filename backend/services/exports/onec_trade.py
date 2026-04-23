"""Export format for 1С:Управление торговлей 11.

Торговля 11 import recognises Код as primary key; Артикул as secondary
search criterion. Has a richer set of columns including Вид номенклатуры,
Производитель, and separate price tables. We keep a pragmatic subset.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .base import ExportFormat, ProductDict


_COLUMNS = [
    ("Код",                 12),
    ("Наименование",        40),
    ("Артикул",             16),
    ("Штрихкод",            18),
    ("Единица измерения",   14),
    ("Группа (папка)",      22),
    ("Вид номенклатуры",    16),
    ("Ставка НДС",          10),
    ("Цена (розничная)",    14),
    ("Цена (закупочная)",   14),
    ("Количество",          12),
]


class OneCTrade11Format(ExportFormat):
    format_id = "onec_trade11"
    label = "1С:Торговля 11"
    description = "Импорт номенклатуры в 1С:Управление торговлей 11"
    target = "onec"

    def generate(self, products: List[ProductDict], meta: Dict[str, Any] | None = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Номенклатура"

        for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, p in enumerate(products, start=2):
            # Code: UT 11 expects an internal code; we leave it empty so that
            # the importer auto-generates one (empty cell is accepted).
            code = self.safe_str(p.get("code") if isinstance(p, dict) else None)
            name = self.safe_str(p.get("name"))
            article = self.safe_str(p.get("article"))
            barcode = self.safe_str(p.get("barcode"))
            unit = self.safe_str(p.get("unit")) or "шт"
            category = self.safe_str(p.get("category"))
            product_type = self.safe_str(p.get("product_type")) or "Товар"
            vat = p.get("vat_rate")
            vat_str = self._format_vat(vat)
            price = self.safe_num(p.get("price"))
            purchase = self.safe_num(p.get("purchase_price"))
            qty = self.safe_num(p.get("quantity"), default=0)

            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=article)
            bc = ws.cell(row=row_idx, column=4, value=barcode)
            bc.number_format = "@"
            ws.cell(row=row_idx, column=5, value=unit)
            ws.cell(row=row_idx, column=6, value=category)
            ws.cell(row=row_idx, column=7, value=product_type)
            ws.cell(row=row_idx, column=8, value=vat_str)
            if price is not None:
                c = ws.cell(row=row_idx, column=9, value=round(price, 2))
                c.number_format = "0.00"
            if purchase is not None:
                c = ws.cell(row=row_idx, column=10, value=round(purchase, 2))
                c.number_format = "0.00"
            if qty is not None:
                c = ws.cell(row=row_idx, column=11, value=round(qty, 3))
                c.number_format = "0.###"

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def _format_vat(value: Any) -> str:
        if value is None or value == "":
            return "НДС 20%"
        try:
            n = int(float(value))
        except (TypeError, ValueError):
            s = str(value).strip()
            return s or "НДС 20%"
        if n <= 0:
            return "Без НДС"
        return f"НДС {n}%"
