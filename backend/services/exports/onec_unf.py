"""Export format for 1С:УНФ (Управление нашей фирмой).

УНФ import wizard uses Артикул (or Код) as the primary key for matching,
with Наименование as a secondary identifier. Column order below mirrors
the default column mapping so the wizard auto-recognises them.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .base import ExportFormat, ProductDict


_COLUMNS = [
    ("Артикул",            16),
    ("Наименование",       40),
    ("Единица",            10),
    ("Группа",             22),
    ("Штрихкод",           18),
    ("Цена продажи",       14),
    ("Цена закупки",       14),
    ("Количество",         12),
    ("Ставка НДС",         10),
]


class OneCUNFFormat(ExportFormat):
    format_id = "onec_unf"
    label = "1С:УНФ"
    description = "Импорт номенклатуры в 1С:Управление нашей фирмой"
    target = "onec"

    def generate(self, products: List[ProductDict], meta: Dict[str, Any] | None = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Номенклатура"

        for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, p in enumerate(products, start=2):
            article = self.safe_str(p.get("article"))
            name = self.safe_str(p.get("name"))
            unit = self.safe_str(p.get("unit")) or "шт"
            category = self.safe_str(p.get("category"))
            barcode = self.safe_str(p.get("barcode"))
            price = self.safe_num(p.get("price"))
            purchase = self.safe_num(p.get("purchase_price"))
            qty = self.safe_num(p.get("quantity"), default=0)
            vat = p.get("vat_rate")
            vat_str = self._format_vat(vat)

            ws.cell(row=row_idx, column=1, value=article)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=unit)
            ws.cell(row=row_idx, column=4, value=category)
            bc = ws.cell(row=row_idx, column=5, value=barcode)
            bc.number_format = "@"
            if price is not None:
                c = ws.cell(row=row_idx, column=6, value=round(price, 2))
                c.number_format = "0.00"
            if purchase is not None:
                c = ws.cell(row=row_idx, column=7, value=round(purchase, 2))
                c.number_format = "0.00"
            if qty is not None:
                c = ws.cell(row=row_idx, column=8, value=round(qty, 3))
                c.number_format = "0.###"
            ws.cell(row=row_idx, column=9, value=vat_str)

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def _format_vat(value: Any) -> str:
        if value is None or value == "":
            return "НДС 20%"
        try:
            n = int(float(value))
        except (TypeError, ValueError):
            s = str(value).strip()
            return s or "НДС 20%"
        if n <= 0:
            return "Без НДС"
        return f"НДС {n}%"
